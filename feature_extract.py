import os
import sys

import cv2
import time
import pickle
import numpy as np
import openpyxl
import openslide
import pyclipper
from scipy.spatial import KDTree, distance_matrix
from skimage import morphology
from PIL import Image
import tqdm
import threading
import multiprocessing


class feature_extraction():
    def __init__(self, save_path):
        self.save_path = save_path
        self.col = ['文件名']
        region_list = ["总", "基质区域", "肿瘤区域"]
        margins = [2000, 1800, 1600, 1400, 1200, 1000, 800, 600, 400, 200, 0, -100, -200, -300, -400]
        for margin in margins:
            if margin > 0:
                region_list.append(f"肿瘤边界{margin}-{margin - 200}um区域")
            else:
                region_list.append(f"肿瘤边界{margin}-{margin - 100}um区域")
        type_list = ["表皮细胞", "淋巴细胞", "中性粒细胞", "邻接的表皮-淋巴细胞", "邻接的表皮-中性粒细胞",
                     "邻接的淋巴-中性粒细胞",
                     "30um内的表皮-淋巴细胞", "30um内的表皮-中性粒细胞", "30um内的淋巴-中性粒细胞",
                     "100um内的表皮-淋巴细胞", "100um内的表皮-中性粒细胞", "100um内的淋巴-中性粒细胞"]
        for type in type_list:
            for region in region_list:
                self.col.append(f"{region}的{type}数")
                if region == '总':
                    continue
                self.col.append(f"{region}的{type}密度")
        self.col.append("表皮-淋巴细胞平均距离")
        self.col.append("表皮-中性粒细胞平均距离")
        self.col.append("淋巴-中性粒细胞平均距离")
        for region in region_list:
            if region == "总":
                continue
            self.col.append(f"{region}面积/um^2")

        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title="特征", index=0)
        for i in range(1, len(self.col) + 1):
            sheet.cell(1, i).value = self.col[i - 1]
        wb.save(save_path)
        self.idx = 2

    def write(self, feature):
        wb = openpyxl.load_workbook(self.save_path)
        sheet = wb['特征']
        for idx, item in enumerate(self.col):
            sheet.cell(self.idx, idx + 1).value = feature[item]
        wb.save(self.save_path)
        self.idx += 1


def get_slide_mpp(slide_path):
    slide = openslide.open_slide(slide_path)
    try:
        mpp = float(slide.properties['openslide.mpp-x'])
    except:
        mpp = 0.24
    return slide, mpp


def get_results(result_dir, slide_path, format):
    slide_name = os.path.basename(slide_path).replace(format, '')

    # region_result_path = os.path.join(result_dir, slide_name, f'{slide_name}_seg.pkl')
    # nuclei_result_path = os.path.join(result_dir, slide_name, f'{slide_name}_nuclei.pkl')

    result_path = os.path.join(result_dir, slide_name, f'{slide_name}.pkl')
    if not os.path.exists(result_path):
        return False, False, False, False, False
    with open(result_path, 'rb') as f:
        result = pickle.load(f)
    f.close()

    # region_result_path1 = os.path.join(result_dir, slide_name, f'{slide_name}.qptiff_seg.pkl')
    # if os.path.exists(region_result_path1):
    #     region_result_path = region_result_path1
    # if not os.path.exists(region_result_path) or not os.path.exists(nuclei_result_path):
    #    return False, False, False, False, False
    # with open(region_result_path, 'rb') as f:
    #    region_result = pickle.load(f)
    # f.close()
    # with open(nuclei_result_path, 'rb') as f:
    #    nuclei_result = pickle.load(f)
    # f.close()

    # 读取结果
    mask = result['mask']
    mask_downsample = result['heatmap_downsample']
    type_arr = result['type']
    center_arr = result['center']
    contour_arr = result['contour']
    # 去除掉之前计算得到的邻接的细胞
    idx = (type_arr != 5) * (type_arr != 6)
    type_arr = type_arr[idx]
    center_arr = center_arr[idx]
    contour_arr = contour_arr[idx]

    return mask, mask_downsample, type_arr, center_arr, contour_arr


# 提取组织区域
def get_tissue_region(slide, mask_level=-1, mthresh=7, use_otsu=False, sthresh=20, sthresh_up=255, close=0):
    if mask_level == -1:
        mask_level = slide.get_best_level_for_downsample(64)
    downsample = slide.level_downsamples[mask_level]
    img = np.array(slide.read_region((0, 0), mask_level, slide.level_dimensions[mask_level]).convert('RGB'))
    img_hsv = cv2.cvtColor(img, cv2.COLOR_RGB2HSV)  # Convert to HSV space
    img_med = cv2.medianBlur(img_hsv[:, :, 1], mthresh)  # Apply median blurring

    # Thresholding
    if use_otsu:
        _, img_otsu = cv2.threshold(img_med, 0, sthresh_up, cv2.THRESH_OTSU + cv2.THRESH_BINARY)
    else:
        _, img_otsu = cv2.threshold(img_med, sthresh, sthresh_up, cv2.THRESH_BINARY)

    # Morphological closing
    if close > 0:
        kernel = np.ones((close, close), np.uint8)
        img_otsu = cv2.morphologyEx(img_otsu, cv2.MORPH_CLOSE, kernel)

    img_otsu = morphology.remove_small_objects(
        img_otsu == 255, min_size=(200 // downsample) ** 2, connectivity=2
    )
    img_otsu = morphology.remove_small_holes(img_otsu, area_threshold=(400 // downsample) ** 2)

    return img_otsu, mask_level


def split_region(slide, mask, mask_downsample, mpp):
    """
            取得肿瘤区域以及其空洞
            对边缘作外扩，对空洞区域作内缩
            先填充外扩的区域，在填充空洞区域
            """
    start = time.time()

    def filter_contours(contours, hierarchy, filter_params: dict):
        """
                        Filter contours by: area.
                    """
        filtered = []
        # find indices of foreground contours (parent == -1)
        hierarchy_1 = np.flatnonzero(hierarchy[:, 1] == -1)
        all_holes = []
        # loop through foreground contour indices
        for cont_idx in hierarchy_1:
            # actual contour
            cont = contours[cont_idx]
            # indices of holes contained in this contour (children of parent contour)
            holes = np.flatnonzero(hierarchy[:, 1] == cont_idx)
            # take contour area (includes holes)
            a = cv2.contourArea(cont)
            # calculate the contour area of each hole
            hole_areas = [cv2.contourArea(contours[hole_idx]) for hole_idx in holes]
            # actual area of foreground contour region
            a = a - np.array(hole_areas).sum()
            if a == 0: continue
            if tuple((filter_params['a_t'],)) < tuple((a,)):
                filtered.append(cont_idx)
                all_holes.append(holes)
        foreground_contours = [contours[cont_idx] for cont_idx in filtered]
        hole_contours = []
        for hole_ids in all_holes:
            unfiltered_holes = [contours[idx] for idx in hole_ids]
            unfilered_holes = sorted(unfiltered_holes, key=cv2.contourArea, reverse=True)
            # take max_n_holes largest holes by area
            unfilered_holes = unfilered_holes[:filter_params['max_n_holes']]
            filtered_holes = []
            # filter these holes
            for hole in unfilered_holes:
                if cv2.contourArea(hole) > filter_params['a_h']:
                    filtered_holes.append(hole)
            hole_contours.append(filtered_holes)
        return foreground_contours, hole_contours

    def equidistant_zoom_contour(contour, margin):
        """
        等距离缩放多边形轮廓点
        :param contour: 一个图形的轮廓格式[[[x1, x2]],...],shape是(-1, 1, 2)
        :param margin: 轮廓外扩的像素距离，margin正数是外扩，负数是缩小
        :return: 外扩后的轮廓点
        """
        margin = int(margin)
        pco = pyclipper.PyclipperOffset()
        ##### 参数限制，默认成2这里设置大一些，主要是用于多边形的尖角是否用圆角代替
        pco.MiterLimit = 10
        contour = contour[::10, 0, :]
        pco.AddPath(contour, pyclipper.JT_MITER, pyclipper.ET_CLOSEDPOLYGON)
        solutions = pco.Execute(margin)
        contour = []
        try:
            for i in range(len(solutions)):
                contour.append(np.array(solutions[i]).reshape(-1, 1, 2).astype(int))
        except:
            contour = np.array([])
        return contour

    tissue_mask, tissue_mask_level = get_tissue_region(slide)
    tissue_mask = cv2.resize(np.array(tissue_mask * 1, dtype=np.uint8), (mask.shape[1], mask.shape[0]),
                             cv2.INTER_NEAREST)
    tissue_mask = np.array(tissue_mask, dtype=bool)
    tissue_mask[mask == 2] = False
    tissue_mask[mask == 3] = False
    mask_tumor = np.uint8((mask == 4) * 255)
    # Image.fromarray(mask_tumor).save('mask/mask_tumor.png')
    # Morphological closing
    # kernel = np.ones((21, 21), np.uint8)
    # mask_tumor = cv2.morphologyEx(mask_tumor, cv2.MORPH_CLOSE, kernel)
    mask_tumor = cv2.GaussianBlur(mask_tumor, (45, 45), 0)
    mask_tumor = (mask_tumor > 128) * 255
    # 过滤掉比较小的区域
    mask_tumor = morphology.remove_small_objects(mask_tumor == 255,
                                                 min_size=(256 / mask_downsample) ** 2, connectivity=2)
    mask_tumor = morphology.remove_small_holes(mask_tumor, area_threshold=(256 / mask_downsample) ** 2)
    mask_tumor = np.array(mask_tumor * 255, dtype=np.uint8)

    # Image.fromarray(mask_tumor).save('mask/mask_tumor1.png')

    # 获取肿瘤边界
    contours, hierarchy = cv2.findContours(mask_tumor, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    hierarchy = np.squeeze(hierarchy, axis=(0,))[:, 2:]
    filter_params = {'a_t': 100, 'a_h': 16, 'max_n_holes': 100}
    foreground_contours, hole_contours = filter_contours(contours, hierarchy, filter_params)

    # 外扩与内缩的距离，来获取不同区域的mask
    margins = [2000, 1800, 1600, 1400, 1200, 1000, 800, 600, 400, 200, -100, -200, -300, -400, -500]
    mask_stroma_regions = {}
    for margin in margins:
        mask_stroma_regions[margin] = np.zeros_like(mask_tumor, dtype=np.uint8)
        for i in range(len(foreground_contours)):
            contour = equidistant_zoom_contour(foreground_contours[i],
                                               margin=(margin / mpp / mask_downsample))
            try:
                cv2.drawContours(mask_stroma_regions[margin], contour, -1, 255, -1)
            except:
                pass
        # 对于空洞区域进行内缩
        for i in range(len(hole_contours)):
            if hole_contours[i] == []:
                continue
            for j in range(len(hole_contours[i])):
                contour = equidistant_zoom_contour(hole_contours[i][j],
                                                   margin=-(margin / mpp / mask_downsample))
                try:
                    cv2.drawContours(mask_stroma_regions[margin], contour, -1, 0, -1)
                except:
                    pass
        # 获得不同层次的区域mask
        mask_stroma_regions[margin] = np.logical_xor(mask_stroma_regions[margin], mask_tumor)
        # 获得组织区域内的mask
        mask_stroma_regions[margin] = np.logical_and(mask_stroma_regions[margin], tissue_mask)
        mask_stroma_regions[margin] = np.array(mask_stroma_regions[margin] * 255, dtype=np.uint8)
        if margin == 2000:
            mask_stroma_regions['基质区域'] = mask_stroma_regions[margin]
    margins = [2000, 1800, 1600, 1400, 1200, 1000, 800, 600, 400]
    for margin in margins:
        mask_stroma_regions[margin] = np.array(
            np.logical_xor(mask_stroma_regions[margin], mask_stroma_regions[margin - 200]) * 255,
            dtype=np.uint8)
    #     Image.fromarray(mask_stroma_regions[margin]).save(f'mask/{margin}.png')
    # Image.fromarray(mask_stroma_regions[200]).save(f'mask/{200}.png')
    margins = [-500, -400, -300, -200]
    for margin in margins:
        mask_stroma_regions[margin] = np.array(
            np.logical_xor(mask_stroma_regions[margin + 100], mask_stroma_regions[margin]) * 255,
            dtype=np.uint8)
        # Image.fromarray(mask_stroma_regions[margin]).save(f'mask/{margin}.png')
    # Image.fromarray(mask_stroma_regions[-100]).save(f'mask/-100.png')
    print(f"划分不同的区域耗费了{time.time() - start:.3f}秒！")

    return mask_tumor, mask_stroma_regions


# 判断两个轮廓之间的最短距离是否小于4
def judge_contours_distance_below4(contour1, contour2):
    distances = distance_matrix(contour1, contour2)
    if distances.min() < 4:
        return True
    else:
        return False


def adjacency_pair(type_arr, center_arr, contour_arr, mpp):
    distance_arr = np.zeros_like(type_arr)
    distance_list = distance_arr.tolist()
    type_list = type_arr.tolist()
    center_list = center_arr.tolist()
    tumor_center_arr, tumor_contour_arr = center_arr[type_arr == 1], contour_arr[type_arr == 1]
    lymph_center_arr, lymph_contour_arr = center_arr[type_arr == 2], contour_arr[type_arr == 2]
    neutrophil_center_arr, neutrophil_contour_arr = center_arr[type_arr == 4], contour_arr[type_arr == 4]

    kdtree_tumor = KDTree(tumor_center_arr)
    kdtree_lymph = KDTree(lymph_center_arr)
    kdtree_neutrophil = KDTree(neutrophil_center_arr)

    # 肿瘤细胞与淋巴细胞
    start = time.time()
    distances_tumor_lymph, points_lymph = kdtree_lymph.query(tumor_center_arr, 1, distance_upper_bound=(100 / mpp))
    distances_tumor_lymph = tqdm.tqdm(distances_tumor_lymph, file=sys.stdout)
    paired_tumor_lymph = {}
    for tumor_idx, distance in enumerate(distances_tumor_lymph):
        if distance == np.inf:
            continue
        lymph_idx = points_lymph[tumor_idx]
        paired_tumor_lymph[tumor_idx] = lymph_idx

        if distance < (30 / mpp):
            # 判断是否邻接
            if judge_contours_distance_below4(tumor_contour_arr[tumor_idx], lymph_contour_arr[lymph_idx]):
                type_list.append(5)
            else:
                type_list.append(6)
        else:
            type_list.append(7)
        distance_list.append(distance)
        center_list.append((tumor_center_arr[tumor_idx] + lymph_center_arr[lymph_idx]) / 2)

    distances_tumor_lymph, points_tumor = kdtree_tumor.query(lymph_center_arr, 1, distance_upper_bound=(100 / mpp))
    distances_tumor_lymph = tqdm.tqdm(distances_tumor_lymph, file=sys.stdout)
    for lymph_idx, distance in enumerate(distances_tumor_lymph):
        if distance == np.inf:
            continue
        tumor_idx = points_tumor[lymph_idx]
        if paired_tumor_lymph.get(tumor_idx) is not None:
            if lymph_idx == paired_tumor_lymph[tumor_idx]:
                continue
        if distance < (30 / mpp):
            # 判断是否邻接
            if judge_contours_distance_below4(tumor_contour_arr[tumor_idx], lymph_contour_arr[lymph_idx]):
                type_list.append(5)
            else:
                type_list.append(6)
        else:
            type_list.append(7)
        distance_list.append(distance)
        center_list.append((tumor_center_arr[tumor_idx] + lymph_center_arr[lymph_idx]) / 2)
    print(f"匹配肿瘤细胞与淋巴细胞耗费了{time.time() - start:.3f}秒！")
    del paired_tumor_lymph, distances_tumor_lymph, points_tumor, points_lymph

    # 肿瘤细胞与中性粒细胞
    start = time.time()
    paired_tumor_neutrophil = {}
    distances_tumor_neutrophil, points_neutrophil = kdtree_neutrophil.query(tumor_center_arr, 1,
                                                                            distance_upper_bound=(100 / mpp))
    distances_tumor_neutrophil = tqdm.tqdm(distances_tumor_neutrophil, file=sys.stdout)
    for tumor_idx, distance in enumerate(distances_tumor_neutrophil):
        if distance == np.inf:
            continue
        neutrophil_idx = points_neutrophil[tumor_idx]
        paired_tumor_neutrophil[tumor_idx] = [neutrophil_idx]
        if distance < (30 / mpp):
            if judge_contours_distance_below4(tumor_contour_arr[tumor_idx], neutrophil_contour_arr[neutrophil_idx]):
                type_list.append(8)
            else:
                type_list.append(9)
        else:
            type_list.append(10)
        distance_list.append(distance)
        center_list.append((tumor_center_arr[tumor_idx] + neutrophil_center_arr[neutrophil_idx]) / 2)
    distances_tumor_neutrophil, points_tumor = kdtree_tumor.query(neutrophil_center_arr, 1,
                                                                  distance_upper_bound=(100 / mpp))
    distances_tumor_neutrophil = tqdm.tqdm(distances_tumor_neutrophil, file=sys.stdout)
    for neutrophil_idx, distance in enumerate(distances_tumor_neutrophil):
        if distance == np.inf:
            continue
        tumor_idx = points_tumor[neutrophil_idx]
        if paired_tumor_neutrophil.get(tumor_idx) is not None:
            if neutrophil_idx == paired_tumor_neutrophil[tumor_idx]:
                continue
        if distance < (30 / mpp):
            if judge_contours_distance_below4(tumor_contour_arr[tumor_idx], neutrophil_contour_arr[neutrophil_idx]):
                type_list.append(8)
            else:
                type_list.append(9)
        else:
            type_list.append(10)
        distance_list.append(distance)
        center_list.append((tumor_center_arr[tumor_idx] + neutrophil_center_arr[neutrophil_idx]) / 2)
    print(f"匹配肿瘤细胞与中性粒细胞耗费了{time.time() - start:.3f}秒！")
    del paired_tumor_neutrophil, distances_tumor_neutrophil, points_neutrophil, points_tumor

    distance_lymph_neutrophil, points_neutrophil = kdtree_neutrophil.query(lymph_center_arr, 1,
                                                                           distance_upper_bound=(100 / mpp))
    distance_lymph_neutrophil = tqdm.tqdm(distance_lymph_neutrophil, file=sys.stdout)
    paired_lymph_neutrophil = {}
    start = time.time()
    for lymph_idx, distance in enumerate(distance_lymph_neutrophil):
        if distance == np.inf:
            continue
        neutrophil_idx = points_neutrophil[lymph_idx]
        paired_lymph_neutrophil[lymph_idx] = neutrophil_idx
        if distance < (30 / mpp):
            if judge_contours_distance_below4(lymph_contour_arr[lymph_idx], neutrophil_contour_arr[neutrophil_idx]):
                type_list.append(11)
            else:
                type_list.append(12)
        else:
            type_list.append(13)
        distance_list.append(distance)
        center_list.append((lymph_center_arr[lymph_idx] + neutrophil_center_arr[neutrophil_idx]) / 2)

    distance_lymph_neutrophil, points_lymph = kdtree_lymph.query(neutrophil_center_arr, 1,
                                                                 distance_upper_bound=(100 / mpp))
    distance_lymph_neutrophil = tqdm.tqdm(distance_lymph_neutrophil, file=sys.stdout)
    for neutrophil_idx, distance in enumerate(distance_lymph_neutrophil):
        if distance == np.inf:
            continue
        lymph_idx = points_lymph[neutrophil_idx]
        if paired_lymph_neutrophil.get(lymph_idx) is not None:
            if neutrophil_idx == paired_lymph_neutrophil[lymph_idx]:
                continue
        if distance < (30 / mpp):
            if judge_contours_distance_below4(lymph_contour_arr[lymph_idx], neutrophil_contour_arr[neutrophil_idx]):
                type_list.append(11)
            else:
                type_list.append(12)
        else:
            type_list.append(13)
        distance_list.append(distance)
        center_list.append((lymph_center_arr[lymph_idx] + neutrophil_center_arr[neutrophil_idx]) / 2)
    print(f"匹配淋巴细胞与中性粒细胞耗费了{time.time() - start:.3f}秒！")
    type_arr = np.array(type_list)
    center_arr = np.array(center_list)
    distance_arr = np.array(distance_list)
    return type_arr, center_arr, distance_arr


def count_nuclei(feature: feature_extraction, slide_name, type_arr, center_arr,
                 mask_tumor, mask_stroma_regions, mask_downsample):
    feature_dict = {}
    for item in feature.col:
        feature_dict[item] = 0
    feature_dict['文件名'] = slide_name
    start = time.time()
    type_dict = {
        1: "表皮细胞",
        2: "淋巴细胞",
        4: "中性粒细胞",
        5: "邻接的表皮-淋巴细胞",
        6: "30um内的表皮-淋巴细胞",
        7: "100um内的表皮-淋巴细胞",
        8: "邻接的表皮-中性粒细胞",
        9: "30um内的表皮-中性粒细胞",
        10: "100um内的表皮-中性粒细胞",
        11: "邻接的淋巴-中性粒细胞",
        12: "30um内的淋巴-中性粒细胞",
        13: "100um内的淋巴-中性粒细胞"
    }
    for type, center in zip(type_arr, center_arr):
        # 跳过巨噬细胞
        if type == 3 or type == 0:
            continue
        type_name = type_dict[type]
        if mask_tumor[int(center[1] // mask_downsample), int(center[0] // mask_downsample)] == 255:
            region_name = "肿瘤区域"
            feature_dict[f"{region_name}的{type_name}数"] += 1
            margins = [-100, -200, -300, -400, -500]
            for margin in margins:
                if mask_stroma_regions[margin][
                    int(center[1] // mask_downsample), int(center[0] // mask_downsample)] == 255:
                    region_name = f"肿瘤边界{margin + 100}-{margin}um区域"
                    feature_dict[f'{region_name}的{type_name}数'] += 1
                    break
        elif mask_stroma_regions['基质区域'][
            int(center[1] // mask_downsample), int(center[0] // mask_downsample)] == 255:
            if type == 1:
                x1 = int((center[1] - 100) // mask_downsample)
                y1 = int((center[0] - 100) // mask_downsample)
                x2 = int((center[1] + 100) // mask_downsample)
                y2 = int((center[0] + 100) // mask_downsample)
                x1 = 0 if x1 < 0 else x1
                y1 = 0 if y1 < 0 else y1
                submask = mask_tumor[x1: x2, y1: y2]
                if submask.sum() == 0:
                    continue
            region_name = "基质区域"
            feature_dict[f"{region_name}的{type_name}数"] += 1
            margins = [200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800, 2000]
            for margin in margins:
                if mask_stroma_regions[margin][
                    int(center[1] // mask_downsample), int(center[0] // mask_downsample)] == 255:
                    region_name = f"肿瘤边界{margin}-{margin - 200}um区域"
                    feature_dict[f'{region_name}的{type_name}数'] += 1
                    break
        else:
            continue
        feature_dict[f'总的{type_name}数'] += 1
    print(f"计算各种细胞在不同区域的数量耗费了{time.time() - start:.3f}秒！")
    return feature_dict


def average_distance_statistics(feature_dict, type_arr, distance_arr, mpp):
    paired_tumor_lymph = (type_arr == 5) + (type_arr == 6) + (type_arr == 7)
    feature_dict["表皮-淋巴细胞平均距离"] = distance_arr[paired_tumor_lymph].sum() / (paired_tumor_lymph.sum()) * mpp
    paired_tumor_neutrophil = (type_arr == 8) + (type_arr == 9) + (type_arr == 10)
    feature_dict["表皮-中性粒细胞平均距离"] = distance_arr[paired_tumor_neutrophil].sum() / (
        paired_tumor_neutrophil.sum()) * mpp
    paired_lymph_neutrophil = (type_arr == 11) + (type_arr == 12) + (type_arr == 13)
    feature_dict["淋巴-中性粒细胞平均距离"] = distance_arr[paired_lymph_neutrophil].sum() / (
        paired_lymph_neutrophil.sum()) * mpp
    return feature_dict


def countArea(feature_dict, mask_tumor, mask_stroma_regions, mask_downsample, mpp):
    # 单位um**2
    feature_dict['肿瘤区域面积/um^2'] = mask_tumor.sum() / 255 * (mpp ** 2) * (mask_downsample ** 2)
    feature_dict['基质区域面积/um^2'] = mask_stroma_regions['基质区域'].sum() / 255 * (mpp ** 2) * (
                mask_downsample ** 2)
    margins = [200, 400, 600, 800, 1000, 1200, 1400, 1600, 1800, 2000]
    for margin in margins:
        feature_dict[f"肿瘤边界{margin}-{margin - 200}um区域面积/um^2"] = mask_stroma_regions[margin].sum() / 255 * (
                    mpp ** 2) * (mask_downsample ** 2)
    margins = [-100, -200, -300, -400, -500]
    for margin in margins:
        feature_dict[f"肿瘤边界{margin + 100}-{margin}um区域面积/um^2"] = mask_stroma_regions[margin].sum() / 255 * (
                    mpp ** 2) * (mask_downsample ** 2)
    return feature_dict


def countDensity(feature_dict):
    region_list = ["基质区域", "肿瘤区域"]
    margins = [2000, 1800, 1600, 1400, 1200, 1000, 800, 600, 400, 200, 0, -100, -200, -300, -400]
    for margin in margins:
        if margin > 0:
            region_list.append(f"肿瘤边界{margin}-{margin - 200}um区域")
        else:
            region_list.append(f"肿瘤边界{margin}-{margin - 100}um区域")
    type_list = ["表皮细胞", "淋巴细胞", "中性粒细胞", "邻接的表皮-淋巴细胞", "邻接的表皮-中性粒细胞",
                 "邻接的淋巴-中性粒细胞",
                 "30um内的表皮-淋巴细胞", "30um内的表皮-中性粒细胞", "30um内的淋巴-中性粒细胞",
                 "100um内的表皮-淋巴细胞", "100um内的表皮-中性粒细胞", "100um内的淋巴-中性粒细胞"]
    for region in region_list:
        for type in type_list:
            if feature_dict[f"{region}面积/um^2"] == 0:
                feature_dict[f"{region}的{type}密度"] = np.nan
            else:
                feature_dict[f"{region}的{type}密度"] = feature_dict[f"{region}的{type}数"] / feature_dict[
                    f"{region}面积/um^2"]
    return feature_dict


if __name__ == '__main__':
    import os
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument('--slide_dir', type=str, default="E:/datasets/test")
    parser.add_argument('--result_dir', type=str, default='results/Microenv')
    parser.add_argument('--save_path', type=str, default='feature.xlsx')
    parser.add_argument('--format', type=str, default='.svs')
    args = parser.parse_args()

    slide_dir = args.slide_dir
    result_dir = args.result_dir
    save_path = args.save_path
    format = args.format
    feature = feature_extraction(save_path)

    abnormal_slide = ["596471-2.svs", "587977-2.svs", "590849.svs"]

    slide_list = os.listdir(slide_dir)
    slide_list = sorted(slide_list)
    for slide_name in slide_list:
        if format not in slide_name:
            continue
        if slide_name in abnormal_slide:
            continue
        print(f"Process {slide_name}")
        start = time.time()
        slide_path = os.path.join(slide_dir, slide_name)
        slide, mpp = get_slide_mpp(slide_path)

        # 从文件中读取数据
        mask, mask_downsample, type_arr, center_arr, contour_arr = get_results(result_dir, slide_path, format)
        if mask is False:
            continue
        # 细分肿瘤边界区域
        mask_tumor, mask_stroma_regions = split_region(slide, mask, mask_downsample, mpp)
        # 计算细胞配对的情况
        type_arr, center_arr, distance_arr = adjacency_pair(type_arr, center_arr, contour_arr, mpp)
        # 数细胞
        feature_dict = count_nuclei(feature, slide_name, type_arr, center_arr,
                                    mask_tumor, mask_stroma_regions, mask_downsample)
        feature_dict = average_distance_statistics(feature_dict, type_arr, distance_arr, mpp)
        feature_dict = countArea(feature_dict, mask_tumor, mask_stroma_regions, mask_downsample, mpp)
        feature_dict = countDensity(feature_dict)
        feature.write(feature_dict)
        print(f"Processing {slide_name} consuming {time.time() - start} seconds!")
